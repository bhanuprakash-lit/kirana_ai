from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Kirana AI Features"

# ── Palette ───────────────────────────────────────────────────────────────────
GREEN_DARK   = "1A7F4E"
GREEN_LIGHT  = "E8F5EE"
BLUE_DARK    = "1A4FA0"
BLUE_LIGHT   = "E8EFFE"
ORANGE_DARK  = "D96B2D"
ORANGE_LIGHT = "FFF3E8"
GREY_DARK    = "3D3D3D"
GREY_MED     = "6B7280"
GREY_LIGHT   = "F3F4F6"
WHITE        = "FFFFFF"
TICK_GREEN   = "16A34A"
CROSS_RED    = "DC2626"
HEADER_BG    = "1E293B"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Segoe UI")

def border_thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    "A": 5,   # #
    "B": 36,  # Feature
    "C": 18,  # Category
    "D": 10,  # Basic
    "E": 10,  # Pro
    "F": 52,  # Limitations / Notes
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ── Title row ─────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 42
ws.merge_cells("A1:F1")
ws["A1"].value    = "🏪  Kirana AI  ·  Feature Matrix"
ws["A1"].fill     = fill(HEADER_BG)
ws["A1"].font     = Font(bold=True, color=WHITE, size=16, name="Segoe UI")
ws["A1"].alignment = center

# ── Sub-header legend ─────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 18
ws.merge_cells("A2:F2")
ws["A2"].value = (
    "✅ = Included   ❌ = Not available   "
    "Basic ₹200/mo · Pro ₹500/mo   |   AI features available on Pro only"
)
ws["A2"].fill      = fill("334155")
ws["A2"].font      = Font(italic=True, color="CBD5E1", size=9, name="Segoe UI")
ws["A2"].alignment = center

# ── Column headers ────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 32
headers = ["#", "Feature", "Category", "Basic", "Pro", "Limitations / Notes"]
header_fills = [HEADER_BG, HEADER_BG, HEADER_BG, "166534", "1D4ED8", HEADER_BG]
for col_idx, (h, bg) in enumerate(zip(headers, header_fills), start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.fill      = fill(bg)
    cell.font      = Font(bold=True, color=WHITE, size=10, name="Segoe UI")
    cell.alignment = center
    cell.border    = border_thin()

# ── Data ──────────────────────────────────────────────────────────────────────
# Format: (Feature, Category, basic, pro, Limitations)
# basic/pro: True = ✅, False = ❌
TICK = "✅"
CROSS = "❌"

rows = [
    # ─ BILLING / POS ──────────────────────────────────────────────────────────
    ("BILLING & POS", None, None, None, None),
    ("Barcode scanning at billing",         "Billing / POS", True,  True,  ""),
    ("Continuous barcode scan mode",        "Billing / POS", True,  True,  ""),
    ("Product search (name, brand, SKU)",   "Billing / POS", True,  True,  ""),
    ("Add to cart & quantity control",      "Billing / POS", True,  True,  ""),
    ("Loose item weight billing",           "Billing / POS", True,  True,  "Enter weight in kg/g; up to 3 decimal places"),
    ("Create & print/share bill",           "Billing / POS", True,  True,  ""),
    ("View today's orders",                 "Billing / POS", True,  True,  ""),
    ("Unknown barcode → add/link product",  "Billing / POS", True,  True,  ""),
    ("AI Campaign bundles shown at POS",    "Billing / POS", True,  True,  "Campaign creation is Pro only"),
    ("User-created Baskets/Bundles at POS", "Billing / POS", True,  True,  "Basket creation requires active subscription"),
    ("Voice Order (AI)",                    "Billing / POS", False, True,  "Pro only · 3 free/day · Max 15 sec recording · Buy credits: 10 for ₹2"),
    ("Handwrite Order (AI)",                "Billing / POS", False, True,  "Pro only · 5 free/day · Auto-detect after 5 s · Buy credits: 10 for ₹3"),

    # ─ STOCK / INVENTORY ──────────────────────────────────────────────────────
    ("STOCK & INVENTORY", None, None, None, None),
    ("View stock levels & product list",    "Stock",         True,  True,  ""),
    ("Add product (manual entry)",          "Stock",         True,  True,  ""),
    ("Add product from global catalog",     "Stock",         True,  True,  "Catalog has 10 000+ pre-loaded products"),
    ("Multi-variant products (size/price)", "Stock",         True,  True,  "Each variant is a separate SKU; unlimited variants"),
    ("Assign barcode to product",           "Stock",         True,  True,  "Scan or type; editable if catalog has no barcode"),
    ("Product categories",                  "Stock",         True,  True,  ""),
    ("Perishable items & expiry tracking",  "Stock",         True,  True,  "Batch-level expiry dates"),
    ("Loose / bulk items",                  "Stock",         True,  True,  "Sold by weight/volume"),
    ("Edit product details & price",        "Stock",         True,  True,  ""),
    ("Low-stock alerts",                    "Stock",         True,  True,  "Threshold-based alerts via in-app & WhatsApp"),
    ("AI stockout prediction",              "Stock",         False, True,  "Pro only · ML-based demand forecasting"),

    # ─ PURCHASE / PROCUREMENT ─────────────────────────────────────────────────
    ("PURCHASE (PROCUREMENT)", None, None, None, None),
    ("Add & manage suppliers",              "Purchase",      False, True,  "Pro only · Phone import from contacts supported"),
    ("Create purchase order",               "Purchase",      False, True,  "Pro only"),
    ("Mark purchase as received",           "Purchase",      False, True,  "Pro only"),
    ("Set payment due dates",               "Purchase",      False, True,  "Pro only"),
    ("Scan Invoice (AI)",                   "Purchase",      False, True,  "Pro only · 2 free/day · Camera, Gallery or PDF · Buy credits: 10 for ₹5"),
    ("Auto-extract line items from bill",   "Purchase",      False, True,  "Pro only · Gemini 2.0 Flash OCR; matches to inventory"),
    ("Confidence score on invoice scan",    "Purchase",      False, True,  "Pro only · Green ≥85%, Orange ≥65%, Red below"),

    # ─ KHATA (UDHAAR) ─────────────────────────────────────────────────────────
    ("KHATA (UDHAAR / CREDIT)", None, None, None, None),
    ("Record customer credit (Udhaar)",     "Khata",         True,  True,  ""),
    ("Collect/recover payment",             "Khata",         True,  True,  "Amount capped at outstanding balance"),
    ("Udhaar statistics dashboard",         "Khata",         True,  True,  "Total due, recovered, overdue counts"),
    ("Customer-wise Udhaar history",        "Khata",         True,  True,  ""),
    ("Add customer from phone contacts",    "Khata",         True,  True,  ""),

    # ─ SUPPLIER PAYMENTS (DISTRIBUTOR KHATA) ──────────────────────────────────
    ("SUPPLIER PAYMENTS", None, None, None, None),
    ("Overdue payment alerts",              "Supplier Pmts", False, True,  "Pro only · Highlighted in red"),
    ("Payments due today",                  "Supplier Pmts", False, True,  "Pro only"),
    ("Upcoming payments (next 7 days)",     "Supplier Pmts", False, True,  "Pro only"),
    ("Paid history (last 7 days)",          "Supplier Pmts", False, True,  "Pro only"),
    ("Mark payment as paid",                "Supplier Pmts", False, True,  "Pro only"),
    ("Payment timeline overview",           "Supplier Pmts", False, True,  "Pro only · Summary totals for each bucket"),

    # ─ CUSTOMERS ──────────────────────────────────────────────────────────────
    ("CUSTOMER MANAGEMENT", None, None, None, None),
    ("Customer profiles (name/phone/email)","Customers",     True,  True,  ""),
    ("Customer segmentation",               "Customers",     True,  True,  "Regular · Occasional · Bulk · Credit · Impulse · Inactive"),
    ("Customer search & filter by segment", "Customers",     True,  True,  ""),
    ("Sync customers from phone contacts",  "Customers",     True,  True,  ""),
    ("Customer purchase history",           "Customers",     True,  True,  ""),
    ("Input validation (name & phone)",     "Customers",     True,  True,  "Blocks XSS chars; phone: 7–15 digits only"),

    # ─ REFERRAL & CAMPAIGNS ───────────────────────────────────────────────────
    ("REFERRAL & CAMPAIGNS", None, None, None, None),
    ("Create referral campaigns",           "Referral",      False, True,  "Pro only · Reward customers for bringing new ones"),
    ("Campaign list & management",          "Referral",      False, True,  "Pro only"),
    ("WhatsApp campaign alerts",            "Referral",      False, True,  "Pro only · Requires WhatsApp Business API setup"),
    ("AI-recommended product campaigns",    "Campaigns",     True,  True,  "Shown on POS home screen"),

    # ─ BASKETS / BUNDLES ──────────────────────────────────────────────────────
    ("BASKETS & BUNDLES", None, None, None, None),
    ("Create product bundles / baskets",    "Baskets",       True,  True,  "Items must exist in inventory"),
    ("Set basket validity (from / to date)","Baskets",       True,  True,  ""),
    ("Delete baskets",                      "Baskets",       True,  True,  ""),
    ("Add basket to cart at POS",           "Baskets",       True,  True,  ""),
    ("WhatsApp alert to customers",         "Baskets",       True,  True,  "Requires WhatsApp Business API; production number needed"),

    # ─ DASHBOARD & ANALYTICS ──────────────────────────────────────────────────
    ("DASHBOARD & ANALYTICS", None, None, None, None),
    ("Business overview / KPIs",            "Dashboard",     True,  True,  "Today's sales, revenue, top products"),
    ("Sales trend charts",                  "Dashboard",     True,  True,  ""),
    ("App usage tracking (admin panel)",    "Dashboard",     True,  True,  "Active/inactive status, daily time-in-app"),
    ("Cashflow / profit reports",           "Dashboard",     False, True,  "Pro only"),
    ("AI demand forecast",                  "Dashboard",     False, True,  "Pro only · ML model per product"),

    # ─ AI CREDITS & BILLING ───────────────────────────────────────────────────
    ("AI CREDITS", None, None, None, None),
    ("Voice credit pack (10 uses)",         "AI Credits",    False, True,  "Pro only · ₹2 per pack · Credits never expire"),
    ("Handwrite credit pack (10 uses)",     "AI Credits",    False, True,  "Pro only · ₹3 per pack · Credits never expire"),
    ("Invoice scan credit pack (10 uses)",  "AI Credits",    False, True,  "Pro only · ₹5 per pack · Credits never expire"),
    ("Daily free AI quota",                 "AI Credits",    False, True,  "Voice 3/day · Handwrite 5/day · Invoice 2/day"),
    ("Credits roll over (no daily expiry)", "AI Credits",    False, True,  "Pro only · Credits valid until used"),

    # ─ SUBSCRIPTION & ACCOUNT ─────────────────────────────────────────────────
    ("SUBSCRIPTION & ACCOUNT", None, None, None, None),
    ("Basic plan",                          "Subscription",  True,  False, "₹200/month (₹7/day) · Core POS + Inventory + Udhaar"),
    ("Pro plan",                            "Subscription",  False, True,  "₹500/month (₹17/day) · All features unlocked"),
    ("Free trial",                          "Subscription",  True,  True,  "Admin-approved · Can be Basic or Pro trial tier"),
    ("In-app subscription management",      "Subscription",  True,  True,  ""),
    ("Multi-account device isolation",      "Subscription",  True,  True,  "AI usage data scoped per user; cleared on logout"),
    ("Push notifications",                  "Subscription",  True,  True,  "Trial expiry alerts, stock-out warnings"),
    ("Quiet hours for notifications",       "Subscription",  True,  True,  "Configurable start/end hour"),
]

# ── Write rows ────────────────────────────────────────────────────────────────
row_num = 4
feature_num = 1
cat_colors = {
    "Billing / POS":  ("EFF6FF", "1D4ED8"),
    "Stock":          ("F0FDF4", "166534"),
    "Purchase":       ("FFF7ED", "9A3412"),
    "Khata":          ("FDF4FF", "7E22CE"),
    "Supplier Pmts":  ("FFF1F2", "9F1239"),
    "Customers":      ("ECFDF5", "065F46"),
    "Referral":       ("FEF9C3", "854D0E"),
    "Campaigns":      ("FEF9C3", "854D0E"),
    "Baskets":        ("F0F9FF", "0C4A6E"),
    "Dashboard":      ("F8FAFC", "1E293B"),
    "AI Credits":     ("FFF7ED", "9A3412"),
    "Subscription":   ("F5F3FF", "4C1D95"),
}

section_headers = {
    "BILLING & POS", "STOCK & INVENTORY", "PURCHASE (PROCUREMENT)",
    "KHATA (UDHAAR / CREDIT)", "SUPPLIER PAYMENTS", "CUSTOMER MANAGEMENT",
    "REFERRAL & CAMPAIGNS", "BASKETS & BUNDLES", "DASHBOARD & ANALYTICS",
    "AI CREDITS", "SUBSCRIPTION & ACCOUNT",
}

for data in rows:
    feature, category, basic, pro, notes = data

    if feature in section_headers:
        # Section header row
        ws.row_dimensions[row_num].height = 22
        ws.merge_cells(f"A{row_num}:F{row_num}")
        cell = ws.cell(row=row_num, column=1, value=f"  {feature}")
        cell.fill      = fill("334155")
        cell.font      = Font(bold=True, color="F1F5F9", size=10, name="Segoe UI")
        cell.alignment = left
        cell.border    = border_thin()
        row_num += 1
        continue

    ws.row_dimensions[row_num].height = 20
    bg, txt = cat_colors.get(category, ("FFFFFF", "000000"))

    # # column
    c = ws.cell(row=row_num, column=1, value=feature_num)
    c.fill = fill(GREY_LIGHT); c.font = font(color=GREY_MED, size=9)
    c.alignment = center; c.border = border_thin()

    # Feature name
    c = ws.cell(row=row_num, column=2, value=feature)
    c.fill = fill(bg)
    c.font = Font(color=GREY_DARK, size=10, name="Segoe UI",
                  bold=(feature_num % 2 == 0))
    c.alignment = left; c.border = border_thin()

    # Category badge
    c = ws.cell(row=row_num, column=3, value=category)
    c.fill = fill(bg)
    c.font = Font(bold=True, color=txt, size=9, name="Segoe UI")
    c.alignment = center; c.border = border_thin()

    # Basic
    c = ws.cell(row=row_num, column=4, value=TICK if basic else CROSS)
    c.fill = fill("DCFCE7" if basic else "FEE2E2")
    c.font = Font(size=13, color=TICK_GREEN if basic else CROSS_RED, name="Segoe UI")
    c.alignment = center; c.border = border_thin()

    # Pro
    c = ws.cell(row=row_num, column=5, value=TICK if pro else CROSS)
    c.fill = fill("DCFCE7" if pro else "FEE2E2")
    c.font = Font(size=13, color=TICK_GREEN if pro else CROSS_RED, name="Segoe UI")
    c.alignment = center; c.border = border_thin()

    # Notes
    c = ws.cell(row=row_num, column=6, value=notes)
    c.fill = fill(bg)
    c.font = Font(color=GREY_MED, size=9, italic=bool(notes), name="Segoe UI")
    c.alignment = left; c.border = border_thin()

    row_num += 1
    feature_num += 1

# ── Freeze panes & auto-filter ────────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{row_num - 1}"

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\Bhanuprakash\Documents\FlutterProjects\kirana_ai\Kirana_AI_Feature_Matrix.xlsx"
wb.save(out)
print(f"Saved: {out}")
